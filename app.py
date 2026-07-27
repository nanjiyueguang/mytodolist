from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from models import db, Todo, TodoStep, StatusHistory, Attachment
from config import Config
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from io import BytesIO
import os
import uuid
import shutil

app = Flask(__name__)
app.config.from_object(Config)

# 禁止缓存静态文件（开发阶段）
@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

db.init_app(app)
with app.app_context():
    db.create_all()

# 确保附件目录存在
os.makedirs(Config.ATTACHMENT_DIR, exist_ok=True)

# ============ 页面路由 ============

@app.route('/')
def index():
    return render_template('index.html')

# ============ Todo API ============

@app.route('/api/todos', methods=['GET'])
def get_todos():
    """获取顶层任务列表（树状）"""
    todos = Todo.query.filter_by(parent_id=None, is_archived=False).order_by(Todo.created_at.desc()).all()
    return jsonify([t.to_dict() for t in todos])

@app.route('/api/todos/archived', methods=['GET'])
def get_archived_todos():
    """获取已归档的顶层任务列表"""
    todos = Todo.query.filter_by(parent_id=None, is_archived=True).order_by(Todo.archived_at.desc()).all()
    return jsonify([t.to_dict() for t in todos])

@app.route('/api/todos', methods=['POST'])
def create_todo():
    """创建任务（可指定 parent_id 创建子任务）"""
    data = request.json
    if not data.get('title'):
        return jsonify({'error': '标题不能为空'}), 400
    
    todo = Todo(
        title=data['title'],
        description=data.get('description', ''),
        priority=data.get('priority', '中'),
        parent_id=data.get('parent_id'),
        start_date=datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else None,
        end_date=datetime.strptime(data['end_date'], '%Y-%m-%d').date() if data.get('end_date') else None,
    )
    db.session.add(todo)
    db.session.flush()
    
    history = StatusHistory(todo_id=todo.id, old_status=None, new_status='待开始', remark='创建任务')
    db.session.add(history)
    db.session.commit()
    
    # 如果是子任务，更新父任务日期
    if todo.parent_id:
        update_parent_dates(todo.id)
    
    return jsonify(todo.to_dict()), 201

@app.route('/api/todos/<int:todo_id>', methods=['GET'])
def get_todo(todo_id):
    """获取单个任务详情"""
    todo = Todo.query.get_or_404(todo_id)
    return jsonify(todo.to_dict())

def update_parent_dates(todo_id):
    """递归向上更新父任务的日期（取子任务的最小开始和最大结束）"""
    todo = Todo.query.get(todo_id)
    if not todo or not todo.parent_id:
        return
    
    parent = Todo.query.get(todo.parent_id)
    if not parent:
        return
    
    # 获取所有子任务的日期（只考虑子任务，不考虑父任务自身）
    children = parent.children.all()
    all_starts = []
    all_ends = []
    
    for child in children:
        if child.start_date:
            all_starts.append(child.start_date)
        if child.end_date:
            all_ends.append(child.end_date)
    
    # 更新父任务日期（完全由子任务决定）
    if all_starts:
        parent.start_date = min(all_starts)
    if all_ends:
        parent.end_date = max(all_ends)
    
    parent.updated_at = datetime.utcnow()
    db.session.commit()
    
    # 递归向上
    update_parent_dates(parent.id)


def _check_all_completed(todo):
    """递归检查任务及所有子任务是否都已完成"""
    if todo.status not in ('已完成', '已取消'):
        return False
    for child in todo.children.all():
        if not _check_all_completed(child):
            return False
    return True

@app.route('/api/todos/<int:todo_id>/archive', methods=['POST'])
def archive_todo(todo_id):
    """归档任务（需主任务及所有子任务都已完成）"""
    todo = Todo.query.get_or_404(todo_id)
    
    # 只允许归档顶层任务
    if todo.parent_id is not None:
        return jsonify({'error': '只能归档顶层主任务'}), 400
    
    # 检查是否所有任务都已完成
    if not _check_all_completed(todo):
        return jsonify({'error': '主任务及所有子任务必须为已完成或已取消状态才能归档'}), 400
    
    # 递归归档主任务及所有子任务
    def archive_recursive(t):
        t.is_archived = True
        t.archived_at = datetime.utcnow()
        for child in t.children.all():
            archive_recursive(child)
    
    archive_recursive(todo)
    db.session.commit()
    
    return jsonify({'message': '归档成功'})

@app.route('/api/todos/<int:todo_id>/restore', methods=['POST'])
def restore_todo(todo_id):
    """还原已归档的任务"""
    todo = Todo.query.get_or_404(todo_id)
    
    if todo.parent_id is not None:
        return jsonify({'error': '只能还原顶层主任务'}), 400
    
    # 递归还原
    def restore_recursive(t):
        t.is_archived = False
        t.archived_at = None
        for child in t.children.all():
            restore_recursive(child)
    
    restore_recursive(todo)
    db.session.commit()
    
    return jsonify({'message': '还原成功'})

# ============ 附件 API ============

@app.route('/api/todos/<int:todo_id>/attachments', methods=['POST'])
def upload_attachment(todo_id):
    """上传附件"""
    todo = Todo.query.get_or_404(todo_id)
    
    if 'file' not in request.files:
        return jsonify({'error': '未找到文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    
    # 检查文件大小
    file.seek(0, 2)
    file_size = file.tell()
    file.seek(0)
    
    if file_size > Config.MAX_ATTACHMENT_SIZE:
        return jsonify({'error': f'文件大小超过限制（最大{Config.MAX_ATTACHMENT_SIZE // 1024 // 1024}MB）'}), 400
    
    # 生成存储文件名
    ext = os.path.splitext(file.filename)[1]
    stored_name = f"{uuid.uuid4().hex}{ext}"
    file_path = os.path.join(Config.ATTACHMENT_DIR, stored_name)
    
    # 保存文件
    file.save(file_path)
    
    # 创建附件记录
    attachment = Attachment(
        todo_id=todo_id,
        filename=file.filename,
        stored_name=stored_name,
        file_size=file_size,
        mime_type=file.content_type or 'application/octet-stream'
    )
    db.session.add(attachment)
    db.session.commit()
    
    return jsonify(attachment.to_dict()), 201

@app.route('/api/todos/<int:todo_id>/attachments', methods=['GET'])
def get_attachments(todo_id):
    """获取任务的附件列表"""
    todo = Todo.query.get_or_404(todo_id)
    attachments = todo.attachments.all()
    return jsonify([a.to_dict() for a in attachments])

@app.route('/api/attachments/<int:attachment_id>/download', methods=['GET'])
def download_attachment(attachment_id):
    """下载附件"""
    attachment = Attachment.query.get_or_404(attachment_id)
    return send_from_directory(
        Config.ATTACHMENT_DIR,
        attachment.stored_name,
        as_attachment=True,
        download_name=attachment.filename
    )

@app.route('/api/attachments/<int:attachment_id>', methods=['DELETE'])
def delete_attachment(attachment_id):
    """删除附件"""
    attachment = Attachment.query.get_or_404(attachment_id)
    
    # 删除文件
    file_path = os.path.join(Config.ATTACHMENT_DIR, attachment.stored_name)
    if os.path.exists(file_path):
        os.remove(file_path)
    
    db.session.delete(attachment)
    db.session.commit()
    
    return jsonify({'message': '删除成功'})

@app.route('/api/todos/<int:todo_id>', methods=['PUT'])
def update_todo(todo_id):
    """更新任务"""
    todo = Todo.query.get_or_404(todo_id)
    data = request.json
    
    todo.title = data.get('title', todo.title)
    todo.description = data.get('description', todo.description)
    todo.priority = data.get('priority', todo.priority)
    
    # 如果进度被设置为100%，自动标记为已完成
    if 'progress' in data and data['progress'] == 100:
        if todo.status not in ('已完成', '已取消'):
            todo.status = '已完成'
            # 触发父任务自动完成检查
            if todo.parent_id:
                _auto_complete_parent(todo.parent_id)
    
    if data.get('start_date'):
        todo.start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
    elif 'start_date' in data:
        todo.start_date = None
    
    if data.get('end_date'):
        todo.end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
    elif 'end_date' in data:
        todo.end_date = None
    
    todo.updated_at = datetime.utcnow()
    db.session.commit()
    
    # 递归更新父任务日期
    update_parent_dates(todo_id)
    
    return jsonify(todo.to_dict())

@app.route('/api/todos/<int:todo_id>/status', methods=['PUT'])
def update_status(todo_id):
    """更新任务状态"""
    todo = Todo.query.get_or_404(todo_id)
    data = request.json
    new_status = data.get('status')
    remark = data.get('remark', '')
    
    if new_status not in Config.STATUS_OPTIONS:
        return jsonify({'error': '无效的状态'}), 400
    
    old_status = todo.status
    todo.status = new_status
    todo.updated_at = datetime.utcnow()
    
    # 完成时自动设置进度100%
    if new_status == '已完成':
        todo.progress = 100
    elif new_status == '已取消':
        todo.progress = 0
    
    history = StatusHistory(todo_id=todo.id, old_status=old_status, new_status=new_status, remark=remark)
    db.session.add(history)
    db.session.commit()
    
    # 检查是否需要自动完成父任务
    if new_status == '已完成' and todo.parent_id:
        _auto_complete_parent(todo.parent_id)
    
    return jsonify(todo.to_dict())


def _auto_complete_parent(parent_id):
    """检查父任务的所有子任务是否都完成，如果是则自动标记父任务为已完成"""
    parent = Todo.query.get(parent_id)
    if not parent:
        return
    
    # 检查所有子任务是否都已完成或已取消
    children = parent.children.all()
    if not children:
        return
    
    all_completed = all(child.status in ('已完成', '已取消') for child in children)
    
    if all_completed and parent.status not in ('已完成', '已取消'):
        old_status = parent.status
        parent.status = '已完成'
        parent.progress = 100
        parent.updated_at = datetime.utcnow()
        
        history = StatusHistory(
            todo_id=parent.id, 
            old_status=old_status, 
            new_status='已完成', 
            remark='所有子任务已完成，自动标记'
        )
        db.session.add(history)
        db.session.commit()
        
        # 递归检查上级父任务
        if parent.parent_id:
            _auto_complete_parent(parent.parent_id)

@app.route('/api/todos/<int:todo_id>', methods=['DELETE'])
def delete_todo(todo_id):
    """删除任务（递归删除子任务）"""
    todo = Todo.query.get_or_404(todo_id)
    _delete_recursive(todo)
    db.session.commit()
    return jsonify({'message': '删除成功'})

def _delete_recursive(todo):
    """递归删除任务及子任务"""
    for child in todo.children.all():
        _delete_recursive(child)
    StatusHistory.query.filter_by(todo_id=todo.id).delete()
    TodoStep.query.filter_by(todo_id=todo.id).delete()
    db.session.delete(todo)

# ============ Step API ============

@app.route('/api/todos/<int:todo_id>/steps', methods=['POST'])
def add_step(todo_id):
    """添加步骤"""
    todo = Todo.query.get_or_404(todo_id)
    data = request.json
    
    if not data.get('title'):
        return jsonify({'error': '步骤标题不能为空'}), 400
    
    max_order = db.session.query(db.func.max(TodoStep.order)).filter_by(todo_id=todo_id).scalar() or 0
    
    step = TodoStep(
        todo_id=todo_id,
        title=data['title'],
        order=max_order + 1
    )
    db.session.add(step)
    db.session.commit()
    return jsonify(step.to_dict()), 201

@app.route('/api/steps/<int:step_id>', methods=['PUT'])
def update_step(step_id):
    """更新步骤（标题/完成状态）"""
    step = TodoStep.query.get_or_404(step_id)
    data = request.json
    
    if 'title' in data:
        step.title = data['title']
    if 'completed' in data:
        step.completed = data['completed']
        step.completed_at = datetime.utcnow() if data['completed'] else None
    if 'order' in data:
        step.order = data['order']
    
    db.session.commit()
    
    return jsonify(step.to_dict())

@app.route('/api/steps/<int:step_id>', methods=['DELETE'])
def delete_step(step_id):
    """删除步骤"""
    step = TodoStep.query.get_or_404(step_id)
    todo_id = step.todo_id
    db.session.delete(step)
    db.session.commit()
    _auto_update_progress(todo_id)
    return jsonify({'message': '删除成功'})

def _auto_update_progress(todo_id):
    """步骤不影响任务进度，此函数保留但不再修改 progress"""
    pass

# ============ 导入导出 API ============

@app.route('/api/todos/template', methods=['GET'])
def download_template():
    """下载导入模板"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "导入模板"
    
    # 表头
    headers = ['任务编号*', '标题*', '描述', '优先级', '状态', '开始日期', '结束日期', '父任务编号']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 示例数据
    sample_data = [
        ['1', '完成项目报告', '撰写Q3季度报告', '高', '进行中', '2026-07-16', '2026-07-20', ''],
        ['1.1', '撰写引言', '项目背景介绍', '中', '已完成', '2026-07-16', '2026-07-17', '1'],
        ['1.2', '数据分析', 'Q3数据整理', '中', '进行中', '2026-07-17', '2026-07-19', '1'],
        ['2', '准备会议材料', '准备周一例会PPT', '中', '待开始', '2026-07-17', '2026-07-18', ''],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # 设置列宽
    for col_letter, width in [('A', 12), ('B', 25), ('C', 30), ('D', 10), ('E', 12), ('F', 12), ('G', 12), ('H', 14)]:
        ws.column_dimensions[col_letter].width = width
    
    # 添加说明工作表
    ws_help = wb.create_sheet("填写说明")
    help_data = [
        ['字段', '必填', '说明'],
        ['任务编号', '是', '唯一标识，如：1, 1.1, 1.2, 2'],
        ['标题', '是', '任务标题，最多200字符'],
        ['描述', '否', '任务描述'],
        ['优先级', '否', '高/中/低，默认为"中"'],
        ['状态', '否', '待开始/进行中/暂挂/已完成/已取消，默认为"待开始"'],
        ['开始日期', '否', '格式：YYYY-MM-DD'],
        ['结束日期', '否', '格式：YYYY-MM-DD'],
        ['父任务编号', '否', '填写父任务的编号，如：1'],
        ['', '', '进度由系统根据步骤自动计算，无需填写'],
    ]
    
    for row, data in enumerate(help_data, 1):
        for col, value in enumerate(data, 1):
            cell = ws_help.cell(row=row, column=col, value=value)
            if row == 1:
                cell.fill = header_fill
                cell.font = header_font
    
    for col_letter, width in [('A', 15), ('B', 8), ('C', 50)]:
        ws_help.column_dimensions[col_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='todo_import_template.xlsx'
    )

@app.route('/api/todos/import', methods=['POST'])
def import_todos():
    """从Excel导入任务"""
    if 'file' not in request.files:
        return jsonify({'error': '未找到文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '文件名为空'}), 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': '仅支持Excel文件(.xlsx/.xls)'}), 400
    
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        # 读取表头
        headers = [cell.value for cell in ws[1]]
        required_fields = ['任务编号*', '标题*']
        
        # 验证必填字段
        for field in required_fields:
            if field not in headers:
                return jsonify({'error': f'缺少必填字段: {field}'}), 400
        
        # 字段映射
        field_map = {
            '任务编号*': 'task_no',
            '标题*': 'title',
            '描述': 'description',
            '优先级': 'priority',
            '状态': 'status',
            '开始日期': 'start_date',
            '结束日期': 'end_date',
            '父任务编号': 'parent_task_no'
        }
        
        # 第一遍：读取所有任务数据
        tasks_data = []
        task_no_map = {}  # 任务编号 -> 索引
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not any(row):
                continue
            
            # 构建数据字典
            data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx] in field_map:
                    field_name = field_map[headers[col_idx]]
                    data[field_name] = value
            
            # 验证必填字段
            if not data.get('task_no'):
                continue
            if not data.get('title'):
                continue
            
            task_no = str(data['task_no']).strip()
            data['task_no'] = task_no
            data['row_idx'] = row_idx
            
            task_no_map[task_no] = len(tasks_data)
            tasks_data.append(data)
        
        # 第二遍：创建任务（先创建父任务，再创建子任务）
        created_todos = {}  # task_no -> todo_id
        success_count = 0
        error_count = 0
        errors = []
        
        def create_todo_with_parent(data):
            nonlocal success_count, error_count
            
            task_no = data['task_no']
            if task_no in created_todos:
                return  # 已创建
            
            # 处理日期
            try:
                if data.get('start_date'):
                    if isinstance(data['start_date'], str):
                        data['start_date'] = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
                    elif hasattr(data['start_date'], 'date'):
                        data['start_date'] = data['start_date'].date()
                
                if data.get('end_date'):
                    if isinstance(data['end_date'], str):
                        data['end_date'] = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
                    elif hasattr(data['end_date'], 'date'):
                        data['end_date'] = data['end_date'].date()
            except ValueError as e:
                errors.append(f"第{data['row_idx']}行: 日期格式错误 - {str(e)}")
                error_count += 1
                return
            
            # 处理优先级
            if data.get('priority') and data['priority'] not in ['高', '中', '低']:
                data['priority'] = '中'
            elif not data.get('priority'):
                data['priority'] = '中'
            
            # 处理状态
            if data.get('status') and data['status'] not in Config.STATUS_OPTIONS:
                data['status'] = '待开始'
            elif not data.get('status'):
                data['status'] = '待开始'
            
            # 处理父任务
            parent_id = None
            parent_task_no = data.get('parent_task_no')
            if parent_task_no:
                parent_task_no = str(parent_task_no).strip()
                if parent_task_no not in task_no_map:
                    errors.append(f"第{data['row_idx']}行: 父任务编号 {parent_task_no} 不存在")
                    error_count += 1
                    return
                
                # 确保父任务已创建
                parent_data = tasks_data[task_no_map[parent_task_no]]
                create_todo_with_parent(parent_data)
                
                if parent_task_no in created_todos:
                    parent_id = created_todos[parent_task_no]
                else:
                    errors.append(f"第{data['row_idx']}行: 父任务创建失败")
                    error_count += 1
                    return
            
            # 创建任务
            try:
                # 根据状态设置初始进度
                initial_progress = 0
                if data['status'] == '已完成':
                    initial_progress = 100
                elif data['status'] == '已取消':
                    initial_progress = 0
                
                todo = Todo(
                    title=str(data['title'])[:200],
                    description=str(data.get('description', '') or ''),
                    priority=data['priority'],
                    status=data['status'],
                    progress=initial_progress,
                    start_date=data.get('start_date'),
                    end_date=data.get('end_date'),
                    parent_id=parent_id
                )
                db.session.add(todo)
                db.session.flush()
                
                # 创建状态历史
                history = StatusHistory(
                    todo_id=todo.id,
                    old_status=None,
                    new_status=data['status'],
                    remark='导入创建'
                )
                db.session.add(history)
                
                created_todos[task_no] = todo.id
                success_count += 1
            except Exception as e:
                errors.append(f"第{data['row_idx']}行: 创建失败 - {str(e)}")
                error_count += 1
        
        # 创建所有任务
        for data in tasks_data:
            create_todo_with_parent(data)
        
        db.session.commit()
        
        result = {
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]  # 最多返回10条错误
        }
        
        return jsonify(result), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'导入失败: {str(e)}'}), 500

# ============ 导出 API ============

@app.route('/api/todos/export', methods=['GET'])
def export_todos():
    """导出todo列表到Excel（按任务层级展示）"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 构建查询（支持日期筛选，不选则全量）
    query = Todo.query.filter_by(parent_id=None)
    if start_date:
        query = query.filter(Todo.created_at >= datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        query = query.filter(Todo.created_at <= datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))
    
    # 获取顶层任务（树状）
    top_todos = query.order_by(Todo.created_at.asc()).all()
    
    # 构建扁平化层级列表: [(level, task_code, parent_code, todo), ...]
    flat_list = []
    
    def flatten_tree(todos, parent_code='', level=0):
        index = 1
        for todo in todos:
            task_code = f"{parent_code}.{index}" if parent_code else str(index)
            flat_list.append((level, task_code, parent_code, todo))
            children = todo.children.order_by(Todo.created_at.asc()).all()
            if children:
                flatten_tree(children, task_code, level + 1)
            index += 1
    
    flatten_tree(top_todos)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Todo列表"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = ['任务编号', '标题', '描述', '优先级', '当前状态', '进度(%)', '开始日期', '结束日期', '步骤(完成/总数)', '父任务编号', '创建时间', '更新时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 层级颜色
    level_fills = {
        0: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        1: PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        2: PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    }
    
    for row_idx, (level, task_code, parent_code, todo) in enumerate(flat_list, 2):
        step_stats = todo.get_step_stats()
        
        # 标题带缩进
        indent = '    ' * level
        title_text = f"{indent}{todo.title}"
        if level > 0:
            title_text = f"{'  ' * level}└ {todo.title}"
        
        ws.cell(row=row_idx, column=1, value=task_code)
        ws.cell(row=row_idx, column=2, value=title_text)
        ws.cell(row=row_idx, column=3, value=todo.description or '')
        ws.cell(row=row_idx, column=4, value=todo.priority)
        ws.cell(row=row_idx, column=5, value=todo.status)
        ws.cell(row=row_idx, column=6, value=todo.progress)
        ws.cell(row=row_idx, column=7, value=todo.start_date.strftime('%Y-%m-%d') if todo.start_date else '')
        ws.cell(row=row_idx, column=8, value=todo.end_date.strftime('%Y-%m-%d') if todo.end_date else '')
        ws.cell(row=row_idx, column=9, value=f"{step_stats['completed']}/{step_stats['total']}")
        ws.cell(row=row_idx, column=10, value=parent_code if parent_code else '')
        ws.cell(row=row_idx, column=11, value=todo.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_idx, column=12, value=todo.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
        
        # 层级背景色
        fill = level_fills.get(level, level_fills[2])
        for col in range(1, 13):
            ws.cell(row=row_idx, column=col).fill = fill
        
        # 顶层任务加粗
        if level == 0:
            bold_font = Font(bold=True)
            ws.cell(row=row_idx, column=1).font = bold_font
            ws.cell(row=row_idx, column=2).font = bold_font
    
    for col_letter, width in [('A',12),('B',40),('C',40),('D',10),('E',12),('F',10),('G',12),('H',12),('I',14),('J',14),('K',20),('L',20)]:
        ws.column_dimensions[col_letter].width = width
    
    # 状态历史工作表
    ws_history = wb.create_sheet("状态变更历史")
    history_headers = ['任务编号', '标题', '原状态', '新状态', '变更时间', '备注']
    for col, header in enumerate(history_headers, 1):
        cell = ws_history.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 构建 id -> task_code 映射
    id_to_code = {}
    for level, task_code, parent_code, todo in flat_list:
        id_to_code[todo.id] = task_code
    
    row = 2
    for level, task_code, parent_code, todo in flat_list:
        for history in todo.status_history.all():
            ws_history.cell(row=row, column=1, value=task_code)
            ws_history.cell(row=row, column=2, value=todo.title)
            ws_history.cell(row=row, column=3, value=history.old_status or '')
            ws_history.cell(row=row, column=4, value=history.new_status)
            ws_history.cell(row=row, column=5, value=history.changed_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws_history.cell(row=row, column=6, value=history.remark or '')
            row += 1
    
    for col_letter, width in [('A',12),('B',30),('C',12),('D',12),('E',20),('F',30)]:
        ws_history.column_dimensions[col_letter].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'todo_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5050, debug=False)
